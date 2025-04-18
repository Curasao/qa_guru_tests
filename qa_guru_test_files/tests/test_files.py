from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from tests.conftest import ARCHIVE_FILE_PATH


def test_pdf_reader():
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zip_file:
        with zip_file.open("test_task.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            assert "DESCRIPTION" in reader.pages[0].extract_text()


def test_xlsx_reader():
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zip_file:
        with zip_file.open("file_xls.xlsx") as xlsx_file:
            reader = load_workbook(xlsx_file)
            assert "1. НАСЕЛЕНИЕ, УЧТЕННОЕ ПРИ \nВСЕРОССИЙСКОЙ ПЕРЕПИСИ НАСЕЛЕНИЯ 2020 ГОДА" == reader.active.cell(row=1, column=1).value


def test_csv_reader():
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zip_file:
        with zip_file.open("test_csv.csv", mode='r') as csv_file:
            reader = csv_file.read().decode("windows-1251", errors="replace")
            assert "Дорожка на стол" in reader